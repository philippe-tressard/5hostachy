"""
Utilitaire d'import des lots depuis un fichier Excel (.xlsx).

Structure attendue du fichier (première ligne = en-tête ignorée) :
  Colonne A : id bâtiment       (entier direct, ex. 1 ; "P" ou vide pour parking)
  Colonne B : numéro de lot     (ex. 102)
  Colonne C : type              (AP, ST, T2, T3, T4, T5, CA, PS, LC, DIV)
  Colonne D : étage             (RDC, 1ER, 2EME, 1SS, 2SS)
  Colonne E : numéro de porte   (optionnel, ignoré)
  Colonne F : N° copropriétaire (ex. 408920)
  Colonne G : Nom copropriétaire (ex. ALIF MASSON)

Règles métier :
  - PS (parking)  : batiment_id = None (le parking n'est pas rattaché à un bâtiment)
  - CA (cave)     : batiment_id = bâtiment du lot résidentiel du même propriétaire
                    (lookup sur le no_coproprietaire parmi les autres lignes)
  - Autres types  : batiment_id résolu depuis la colonne A

Chaque ligne importée crée un enregistrement staging ``LotImport``.
L'admin résout ensuite chaque ligne dans l'interface /admin/lots-import.

Usage depuis le conteneur :
  docker compose exec api python -c "
  from app.utils.import_lots import importer_depuis_fichier
  importer_depuis_fichier('/import-data/213 - liste des lots.xlsx')
  "
"""
from __future__ import annotations

import io
import unicodedata
from datetime import datetime
from pathlib import Path
from typing import Optional

from sqlmodel import Session, select

from app.database import engine
from app.models.core import LotImport, StatutLotImport


# -- Types de lots spéciaux ---------------------------------------------------

def _est_parking(type_raw: Optional[str]) -> bool:
    """Vrai si le lot est un parking (PS)."""
    return bool(type_raw and type_raw.strip().upper().startswith("PS"))


def _est_cave(type_raw: Optional[str]) -> bool:
    """Vrai si le lot est une cave (CA)."""
    return bool(type_raw and type_raw.strip().upper().startswith("CA"))


# -- Helpers -------------------------------------------------------------------

def normaliser(s: Optional[str]) -> str:
    """Normalise une chaîne : majuscules, sans accents, espaces normalisés."""
    if not s:
        return ""
    s = s.strip().upper()
    s = "".join(
        c for c in unicodedata.normalize("NFD", s)
        if unicodedata.category(c) != "Mn"
    )
    return " ".join(s.split())


def _resolve_batiment_id(bat_raw) -> Optional[int]:
    """Résout la valeur de la colonne A en id bâtiment.
    Retourne None si la valeur n'est pas un entier valide (ex. 'P', vide)."""
    if bat_raw is None:
        return None
    if isinstance(bat_raw, int):
        return bat_raw
    if isinstance(bat_raw, float) and bat_raw.is_integer():
        return int(bat_raw)
    # Tenter une conversion depuis string
    s = str(bat_raw).strip()
    if s.isdigit():
        return int(s)
    return None


# -- Import en base (staging) --------------------------------------------------

def importer_depuis_bytes(
    contenu: bytes,
    session: Session,
    remplacer: bool = False,
) -> dict:
    """Import depuis des bytes en mémoire (endpoint upload FastAPI)."""
    try:
        import openpyxl  # type: ignore
    except ImportError:
        raise RuntimeError("openpyxl n'est pas installé.")
    wb = openpyxl.load_workbook(io.BytesIO(contenu), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    stats = _traiter_rows(rows, session, remplacer)
    session.commit()
    return stats


def importer_depuis_fichier(chemin: str, remplacer: bool = False) -> dict:
    """
    Importe les lots depuis un xlsx (script CLI dans le conteneur).

    Args:
        chemin    : chemin absolu vers le fichier Excel.
        remplacer : si True, supprime les enregistrements ``en_attente``
                    existants avant de ré-importer.

    Returns:
        dict avec les clés ``importes``, ``ignores``, ``doublons``, ``erreurs``.
    """
    try:
        import openpyxl  # type: ignore
    except ImportError:
        raise RuntimeError("openpyxl n'est pas installé. Ajouter au requirements.txt.")

    path = Path(chemin)
    if not path.exists():
        raise FileNotFoundError(f"Fichier introuvable : {chemin}")

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    with Session(engine) as session:
        stats = _traiter_rows(rows, session, remplacer)
        session.commit()
    return stats


def _traiter_rows(rows: list, session: Session, remplacer: bool) -> dict:
    """Traite les lignes du classeur et crée les LotImport en staging.

    Deux passes :
    - Passe 1 : construit un mapping no_coproprietaire → batiment_id
      à partir des lots résidentiels (ni cave ni parking).
    - Passe 2 : importe toutes les lignes :
        PS  → batiment_id = None
        CA  → batiment_id du lot résidentiel du même propriétaire (passe 1)
        autres → batiment_id résolu depuis la colonne A
    """
    data_rows = rows[1:]  # ignorer l'en-tête
    stats: dict = {"importes": 0, "ignores": 0, "doublons": 0, "erreurs": []}

    if remplacer:
        existants = session.exec(
            select(LotImport).where(
                LotImport.statut != StatutLotImport.resolu
            )
        ).all()
        for e in existants:
            session.delete(e)
        session.flush()

    # ── Passe 1 : mapping propriétaire → bâtiment (lots résidentiels) ──────
    owner_bat: dict[str, int] = {}  # no_coproprietaire → batiment_id
    for row in data_rows:
        if not row or len(row) < 3:
            continue
        t = str(row[2]).strip() if row[2] else None
        if not t or _est_cave(t) or _est_parking(t):
            continue
        no = str(row[5]).strip() if len(row) > 5 and row[5] not in (None, "") else None
        if no and no not in owner_bat:
            bid = _resolve_batiment_id(row[0])
            if bid is not None:
                owner_bat[no] = bid

    # ── Passe 2 : import ────────────────────────────────────────────────────
    for i, row in enumerate(data_rows, start=2):
        if not row or len(row) < 3:
            continue

        bat_raw   = row[0]
        num_raw   = str(row[1]).strip() if row[1] is not None else None
        type_raw  = str(row[2]).strip() if row[2] else None
        etage_raw = str(row[3]).strip() if len(row) > 3 and row[3] else None
        no_cop    = str(row[5]).strip() if len(row) > 5 and row[5] not in (None, "") else None
        nom_cop   = str(row[6]).strip() if len(row) > 6 and row[6] not in (None, "") else None

        # Champs requis (bat_raw peut être absent pour parking)
        if not num_raw or not type_raw:
            stats["ignores"] += 1
            continue

        # ── Résolution du bâtiment selon le type ───────────────────────────
        if _est_parking(type_raw):
            bat_id = None          # parking : pas de bâtiment
        elif _est_cave(type_raw):
            # Cave : bâtiment = bâtiment du lot résidentiel du même propriétaire
            bat_id = owner_bat.get(no_cop) if no_cop else None
            if bat_id is None:
                # Fallback : tenter la résolution directe depuis la colonne A
                bat_id = _resolve_batiment_id(bat_raw)
        else:
            bat_id = _resolve_batiment_id(bat_raw)
            if bat_id is None:
                stats["erreurs"].append(
                    f"Ligne {i} — bâtiment inconnu : {bat_raw!r} (lot {num_raw})"
                )
                continue

        numero = normaliser(num_raw)

        # ── Dédoublonnage en staging ────────────────────────────────────────
        doublon_q = select(LotImport).where(LotImport.numero == numero)
        if bat_id is not None:
            doublon_q = doublon_q.where(LotImport.batiment_id == bat_id)
        else:
            doublon_q = doublon_q.where(LotImport.batiment_id.is_(None))  # type: ignore[union-attr]
        if session.exec(doublon_q).first():
            stats["doublons"] += 1
            continue

        session.add(LotImport(
            batiment_id=bat_id,
            numero=numero,
            type_raw=type_raw,
            etage_raw=etage_raw,
            no_coproprietaire=no_cop,
            nom_coproprietaire=nom_cop,
            importe_le=datetime.utcnow(),
        ))
        stats["importes"] += 1

    return stats
