"""
Utilitaire d'import des télécommandes depuis un fichier Excel (.xlsx).

Structure attendue du fichier (première ligne = en-tête ignorée) :
  Colonne A : nom du copropriétaire (requis)
  Colonne B : nom du locataire (optionnel)
  Colonne C : référence de la télécommande (optionnel sur quelques lignes spéciales)

Usage depuis le conteneur :
  docker compose exec api python -c "
  from app.utils.import_telecommandes import importer_depuis_fichier
  importer_depuis_fichier('/chemin/vers/fichier.xlsx')
  "
"""
from __future__ import annotations

import io
import unicodedata
from datetime import datetime
from pathlib import Path
from typing import Optional

from sqlmodel import Session, select

from app.database import engine
from app.models.core import StatutImport, TelecommandeImport


# ── Noms à ignorer automatiquement (accès non-résidents) ──────────────────
_NOMS_IGNORES = {
    "PARKINGS PUBLIQUES",
    "0. ACCES MAIRIE",
    "ATPE",
    "CDE 22/06/2023",
    "- CDE 02/07/2024",
}


def normaliser(s: Optional[str]) -> str:
    """Normalise une chaîne : majuscules, sans accents, espaces normalisés."""
    if not s:
        return ""
    s = s.strip().upper()
    s = "".join(
        c for c in unicodedata.normalize("NFD", s)
        if unicodedata.category(c) != "Mn"
    )
    return " ".join(s.split())


def importer_depuis_bytes(
    contenu: bytes,
    session: Session,
    remplacer: bool = False,
) -> dict:
    """Import depuis des bytes en mémoire (endpoint upload FastAPI)."""
    try:
        import openpyxl  # type: ignore
    except ImportError:
        raise RuntimeError("openpyxl n'est pas installé.")
    wb = openpyxl.load_workbook(io.BytesIO(contenu), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    stats = _traiter_rows(rows, session, remplacer)
    session.commit()
    return stats


def importer_depuis_fichier(chemin: str, remplacer: bool = False) -> dict:
    """
    Importe les télécommandes depuis un xlsx (script CLI dans le conteneur).

    Args:
        chemin    : chemin absolu vers le fichier Excel.
        remplacer : si True, supprime tous les enregistrements `en_attente`
                    existants avant de ré-importer.

    Returns:
        dict avec les clés `importes`, `ignores`, `doublons`, `erreurs`.
    """
    try:
        import openpyxl  # type: ignore
    except ImportError:
        raise RuntimeError("openpyxl n'est pas installé. Ajouter au requirements.txt.")

    path = Path(chemin)
    if not path.exists():
        raise FileNotFoundError(f"Fichier introuvable : {chemin}")

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    with Session(engine) as session:
        stats = _traiter_rows(rows, session, remplacer)
        session.commit()
    return stats


def _traiter_rows(rows: list, session: Session, remplacer: bool) -> dict:
    """Traite les lignes du classeur et insère les TelecommandeImport."""
    data_rows = rows[1:]  # ignorer l'en-tête
    stats: dict = {"importes": 0, "ignores": 0, "doublons": 0, "erreurs": []}

    if remplacer:
        existants = session.exec(
            select(TelecommandeImport).where(
                TelecommandeImport.statut == StatutImport.en_attente
            )
        ).all()
        for e in existants:
            session.delete(e)
        session.flush()

    for i, row in enumerate(data_rows, start=2):
        nom_prop_raw  = str(row[0]).strip() if row[0] else None
        nom_loc_raw   = str(row[1]).strip() if row[1] else None
        reference_raw = str(row[2]).strip() if row[2] else None

        if not nom_prop_raw:
            continue

        nom_prop_norm = normaliser(nom_prop_raw)

        if nom_prop_norm in _NOMS_IGNORES:
            stats["ignores"] += 1
            _creer_import(
                session, nom_prop_raw, nom_loc_raw, reference_raw,
                statut=StatutImport.ignore,
                notes_admin=f"Ignoré automatiquement (hors résidents) — ligne Excel {i}",
            )
            continue

        if nom_loc_raw and normaliser(nom_loc_raw) == nom_prop_norm:
            nom_loc_raw = None  # proprio-occupant

        if reference_raw:
            doublon = session.exec(
                select(TelecommandeImport).where(
                    TelecommandeImport.nom_proprietaire == nom_prop_raw,
                    TelecommandeImport.reference == reference_raw,
                )
            ).first()
            if doublon:
                stats["doublons"] += 1
                continue

        _creer_import(session, nom_prop_raw, nom_loc_raw, reference_raw)
        stats["importes"] += 1

    return stats


def _creer_import(
    session: Session,
    nom_proprietaire: str,
    nom_locataire: Optional[str],
    reference: Optional[str],
    statut: StatutImport = StatutImport.en_attente,
    notes_admin: Optional[str] = None,
) -> TelecommandeImport:
    record = TelecommandeImport(
        nom_proprietaire=nom_proprietaire,
        nom_locataire=nom_locataire or None,
        reference=reference or None,
        statut=statut,
        notes_admin=notes_admin,
        importe_le=datetime.utcnow(),
    )
    session.add(record)
    return record
