"""
Utilitaire d'import des vigiks depuis un fichier Excel (.xlsx).

Structure attendue du fichier (première ligne = en-tête ignorée) :
  Colonne A : numéro de bâtiment   (ex. « 3 »)
  Colonne B : numéro d'appartement (ex. « 102 »)
  Colonne C : nom du copropriétaire (requis)
  Colonne D : nom du locataire (optionnel)
  Colonne E : N° CLÉS — référence du vigik

Usage depuis le conteneur :
  docker compose exec api python -c "
  from app.utils.import_vigiks import importer_depuis_fichier
  importer_depuis_fichier('/chemin/vers/vigik.xlsx')
  "
"""
from __future__ import annotations

import io
import unicodedata
from datetime import datetime
from pathlib import Path
from typing import Optional

from sqlmodel import Session, select

from app.database import engine
from app.models.core import Batiment, Lot, StatutImport, VigikImport


# ── Noms à ignorer automatiquement ────────────────────────────────────────
_NOMS_IGNORES = {
    "PARKINGS PUBLIQUES",
    "0. ACCES MAIRIE",
    "ATPE",
}


def normaliser(s: Optional[str]) -> str:
    """Normalise une chaîne : majuscules, sans accents, espaces normalisés."""
    if not s:
        return ""
    s = s.strip().upper()
    s = "".join(
        c for c in unicodedata.normalize("NFD", s)
        if unicodedata.category(c) != "Mn"
    )
    return " ".join(s.split())


def importer_depuis_bytes(
    contenu: bytes,
    session: Session,
    remplacer: bool = False,
) -> dict:
    """Import depuis des bytes en mémoire (endpoint upload FastAPI)."""
    try:
        import openpyxl  # type: ignore
    except ImportError:
        raise RuntimeError("openpyxl n'est pas installé.")
    wb = openpyxl.load_workbook(io.BytesIO(contenu), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    stats = _traiter_rows(rows, session, remplacer)
    session.commit()
    return stats


def importer_depuis_fichier(chemin: str, remplacer: bool = False) -> dict:
    """
    Importe les vigiks depuis un xlsx (script CLI dans le conteneur).

    Args:
        chemin    : chemin absolu vers le fichier Excel.
        remplacer : si True, supprime tous les enregistrements `en_attente`
                    existants avant de ré-importer.

    Returns:
        dict avec les clés `importes`, `ignores`, `doublons`, `erreurs`.
    """
    try:
        import openpyxl  # type: ignore
    except ImportError:
        raise RuntimeError("openpyxl n'est pas installé. Ajouter au requirements.txt.")

    path = Path(chemin)
    if not path.exists():
        raise FileNotFoundError(f"Fichier introuvable : {chemin}")

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    with Session(engine) as session:
        stats = _traiter_rows(rows, session, remplacer)
        session.commit()
    return stats


def _traiter_rows(rows: list, session: Session, remplacer: bool) -> dict:
    """Traite les lignes du classeur et insère les VigikImport."""
    data_rows = rows[1:]  # ignorer l'en-tête
    stats: dict = {"importes": 0, "ignores": 0, "doublons": 0, "erreurs": []}

    if remplacer:
        existants = session.exec(
            select(VigikImport).where(
                VigikImport.statut == StatutImport.en_attente
            )
        ).all()
        for e in existants:
            session.delete(e)
        session.flush()

    # Pré-charger l'index lot (batiment_numero, lot_numero) → lot_id
    lot_index = _build_lot_index(session)

    for i, row in enumerate(data_rows, start=2):
        if len(row) < 5:
            continue

        batiment_raw  = str(row[0]).strip() if row[0] else None
        appt_raw      = str(row[1]).strip() if row[1] else None
        nom_prop_raw  = str(row[2]).strip() if row[2] else None
        nom_loc_raw   = str(row[3]).strip() if row[3] else None
        code_raw      = str(row[4]).strip() if row[4] else None

        if not nom_prop_raw:
            continue

        nom_prop_norm = normaliser(nom_prop_raw)

        if nom_prop_norm in _NOMS_IGNORES:
            stats["ignores"] += 1
            _creer_import(
                session, batiment_raw, appt_raw, nom_prop_raw, nom_loc_raw, code_raw,
                statut=StatutImport.ignore,
                notes_admin=f"Ignoré automatiquement (hors résidents) — ligne Excel {i}",
            )
            continue

        # Déduplique propriétaire-occupant
        if nom_loc_raw and normaliser(nom_loc_raw) == nom_prop_norm:
            nom_loc_raw = None

        # Éviter les doublons exacts (même propriétaire + même code)
        if code_raw:
            doublon = session.exec(
                select(VigikImport).where(
                    VigikImport.nom_proprietaire == nom_prop_raw,
                    VigikImport.code == code_raw,
                )
            ).first()
            if doublon:
                stats["doublons"] += 1
                continue

        # Tentative de résolution automatique du lot
        lot_id: Optional[int] = None
        if batiment_raw and appt_raw:
            lot_id = lot_index.get((normaliser(batiment_raw), normaliser(appt_raw)))

        _creer_import(session, batiment_raw, appt_raw, nom_prop_raw, nom_loc_raw, code_raw, lot_id=lot_id)
        stats["importes"] += 1

    return stats


def _build_lot_index(session: Session) -> dict[tuple[str, str], int]:
    """Construit un dictionnaire (batiment_num_norm, lot_num_norm) → lot_id."""
    index: dict[tuple[str, str], int] = {}
    batiments = session.exec(select(Batiment)).all()
    bat_map = {b.id: normaliser(b.numero) for b in batiments}
    lots = session.exec(select(Lot)).all()
    for lot in lots:
        bat_num = bat_map.get(lot.batiment_id, "")
        lot_num = normaliser(lot.numero)
        if bat_num and lot_num:
            index[(bat_num, lot_num)] = lot.id
    return index


def _creer_import(
    session: Session,
    batiment_raw: Optional[str],
    appartement_raw: Optional[str],
    nom_proprietaire: str,
    nom_locataire: Optional[str],
    code: Optional[str],
    lot_id: Optional[int] = None,
    statut: StatutImport = StatutImport.en_attente,
    notes_admin: Optional[str] = None,
) -> VigikImport:
    record = VigikImport(
        batiment_raw=batiment_raw,
        appartement_raw=appartement_raw,
        nom_proprietaire=nom_proprietaire,
        nom_locataire=nom_locataire or None,
        code=code or None,
        lot_id=lot_id,
        statut=statut,
        notes_admin=notes_admin,
        importe_le=datetime.utcnow(),
    )
    session.add(record)
    return record
